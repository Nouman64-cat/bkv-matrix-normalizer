"""Excel file processor for converting XLSX files to structured data."""

import pandas as pd
import openpyxl
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
from io import BytesIO
import logging
from datetime import datetime, timezone

from app.core.config import get_settings
from app.core.exceptions import FileProcessingError

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Processes Excel files and converts them to structured data."""

    def __init__(self):
        self.settings = get_settings()

    def process_file(self, file_content: bytes, filename: str) -> Dict[str, Any]:
        """
        Process Excel file and extract structured data.

        Args:
            file_content: Raw file content as bytes
            filename: Original filename

        Returns:
            Dictionary containing processed data and metadata
        """
        try:
            logger.info(f"Processing Excel file: {filename}")

            # Load workbook
            workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)

            # Get all sheet names
            sheet_names = workbook.sheetnames
            logger.info(f"Found {len(sheet_names)} sheets: {sheet_names}")

            # Process each sheet
            processed_data = {}
            for sheet_name in sheet_names:
                try:
                    sheet_data = self._process_sheet(workbook, sheet_name)
                    if sheet_data:  # Only include non-empty sheets
                        processed_data[sheet_name] = sheet_data
                except Exception as e:
                    logger.warning(f"Failed to process sheet '{sheet_name}': {str(e)}")
                    continue

            workbook.close()

            if not processed_data:
                raise FileProcessingError("No valid data found in Excel file")

            # Return structured result
            return {
                "filename": filename,
                "file_type": "xlsx",
                "sheets": processed_data,
                "sheet_count": len(processed_data),
                "processed_at": datetime.now(timezone.utc).isoformat(),
                "total_rows": sum(
                    len(sheet_data.get("data", []))
                    for sheet_data in processed_data.values()
                ),
            }

        except Exception as e:
            logger.error(f"Failed to process Excel file {filename}: {str(e)}")
            raise FileProcessingError(f"Excel processing failed: {str(e)}")

    def _process_sheet(
        self, workbook: openpyxl.Workbook, sheet_name: str
    ) -> Optional[Dict[str, Any]]:
        """
        Process individual Excel sheet.

        Args:
            workbook: Openpyxl workbook object
            sheet_name: Name of the sheet to process

        Returns:
            Dictionary containing sheet data and metadata
        """
        try:
            worksheet = workbook[sheet_name]

            # Convert to pandas DataFrame for easier processing
            data = []
            headers = []

            # Get all rows with data
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                return None

            # Remove completely empty rows
            rows = [row for row in rows if any(cell is not None for cell in row)]
            if not rows:
                return None

            # First row as headers
            headers = [
                str(cell) if cell is not None else f"Column_{i+1}"
                for i, cell in enumerate(rows[0])
            ]

            # Process data rows
            for row in rows[1:]:
                row_data = {}
                for i, cell in enumerate(row):
                    if i < len(headers):
                        # Convert cell value to appropriate type
                        processed_value = self._process_cell_value(cell)
                        row_data[headers[i]] = processed_value

                # Only add row if it has at least one non-null value
                if any(value is not None for value in row_data.values()):
                    data.append(row_data)

            if not data:
                return None

            return {
                "headers": headers,
                "data": data,
                "row_count": len(data),
                "column_count": len(headers),
                "sheet_name": sheet_name,
            }

        except Exception as e:
            logger.error(f"Failed to process sheet {sheet_name}: {str(e)}")
            return None

    def _process_cell_value(self, cell_value: Any) -> Any:
        """
        Process individual cell value and convert to appropriate Python type.

        Args:
            cell_value: Raw cell value from openpyxl

        Returns:
            Processed cell value
        """
        if cell_value is None:
            return None

        # Handle datetime objects
        if isinstance(cell_value, datetime):
            return cell_value.isoformat()

        # Handle numeric values
        if isinstance(cell_value, (int, float)):
            # Check if it's a whole number that can be represented as int
            if isinstance(cell_value, float) and cell_value.is_integer():
                return int(cell_value)
            return cell_value

        # Handle boolean values
        if isinstance(cell_value, bool):
            return cell_value

        # Handle string values
        if isinstance(cell_value, str):
            # Clean up whitespace
            cleaned = cell_value.strip()

            # Try to convert to number if it looks like one
            if cleaned.replace(".", "").replace("-", "").replace("+", "").isdigit():
                try:
                    if "." in cleaned:
                        float_val = float(cleaned)
                        if float_val.is_integer():
                            return int(float_val)
                        return float_val
                    else:
                        return int(cleaned)
                except ValueError:
                    pass

            # Try to convert to boolean
            if cleaned.lower() in ("true", "false", "yes", "no", "1", "0"):
                return cleaned.lower() in ("true", "yes", "1")

            return cleaned

        # Return as string for any other type
        return str(cell_value)

    def get_preview(
        self, file_content: bytes, filename: str, max_rows: int = None
    ) -> Dict[str, Any]:
        """
        Get a preview of the Excel file data.

        Args:
            file_content: Raw file content as bytes
            filename: Original filename
            max_rows: Maximum number of rows to preview per sheet

        Returns:
            Dictionary containing preview data
        """
        max_rows = max_rows or self.settings.MAX_ROWS_PREVIEW

        try:
            logger.info(f"Generating preview for Excel file: {filename}")

            workbook = openpyxl.load_workbook(BytesIO(file_content), data_only=True)
            sheet_names = workbook.sheetnames

            preview_data = {}
            for sheet_name in sheet_names[:3]:  # Preview max 3 sheets
                try:
                    worksheet = workbook[sheet_name]
                    rows = list(worksheet.iter_rows(values_only=True))

                    if not rows:
                        continue

                    # Get headers
                    headers = [
                        str(cell) if cell is not None else f"Column_{i+1}"
                        for i, cell in enumerate(rows[0])
                    ]

                    # Get preview rows
                    preview_rows = []
                    for row in rows[1 : max_rows + 1]:
                        row_data = {}
                        for i, cell in enumerate(row):
                            if i < len(headers):
                                row_data[headers[i]] = self._process_cell_value(cell)
                        preview_rows.append(row_data)

                    preview_data[sheet_name] = {
                        "headers": headers,
                        "preview_data": preview_rows,
                        "total_rows": len(rows) - 1,  # Excluding header
                        "preview_rows": len(preview_rows),
                    }

                except Exception as e:
                    logger.warning(f"Failed to preview sheet '{sheet_name}': {str(e)}")
                    continue

            workbook.close()

            return {
                "filename": filename,
                "file_type": "xlsx",
                "sheets": preview_data,
                "sheet_count": len(preview_data),
            }

        except Exception as e:
            logger.error(f"Failed to generate preview for {filename}: {str(e)}")
            raise FileProcessingError(f"Preview generation failed: {str(e)}")


# Global processor instance
excel_processor = ExcelProcessor()
